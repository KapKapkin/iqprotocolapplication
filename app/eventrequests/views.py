from copy import copy
from django.http import FileResponse, HttpResponseRedirect, Http404
from django.shortcuts import render
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.templatetags.static import static
from django.core.files.storage import default_storage
from django.conf import settings

import os
import openpyxl

from .models import Event, AvWindow, HonoredGuest, Subceremony, Speaker
from .forms import EventForm, HonoredGuestFormset, SubceremonyFormset, SpeakerFormset

from django.core.mail import EmailMessage


def send_email(email, link):
    email = EmailMessage(
        'Вы отправили заявку на новое событие.', 'Здравствуйте!\n Вы получили это письмо, потому что с вашего аккаунта была отправлена заявка на протокольное сопровождение мероприятия. %s \nС уважением, IQProtocol!' % link, to=[email])
    email.send()


@login_required
def delete_window(request, *args, **kwargs):
    if request.user.is_staff or request.user.is_superuser:
        object = AvWindow.objects.get(id=kwargs['pk'])
        object.delete()
        return HttpResponseRedirect(reverse('account'))
    else:
        return Http404()


@login_required
def create_event(request):
    template_name = 'form.html'
    if request.method == 'GET':
        event_form = EventForm()
        honoredguests_formset = HonoredGuestFormset(
            request.GET or None, queryset=Event.objects.none(), prefix="guests")
        subceremony_formset = SubceremonyFormset(
            request.GET or None, queryset=Event.objects.none(), prefix="subceremony")
        speakers_formset = SpeakerFormset(
            request.GET or None, queryset=Event.objects.none(), prefix="speakers")
    else:
        event_form = EventForm(request.POST)
        honoredguests_formset = HonoredGuestFormset(
            request.POST, prefix="guests")
        subceremony_formset = SubceremonyFormset(
            request.POST, prefix="subceremony")
        speakers_formset = SpeakerFormset(request.POST, prefix="speakers")
        if event_form.is_valid():
            event = event_form.save(commit=False)
            event.user = request.user
            subceremony_formset = SubceremonyFormset(
                request.POST, prefix="subceremony", instance=event)
            if event_form.is_valid() and subceremony_formset.is_valid() and honoredguests_formset.is_valid() and speakers_formset.is_valid():
                try:
                    event.save()
                    subceremony_formset.save()

                    for instance in honoredguests_formset:
                        instance = instance.save(commit=False)
                        instance.event = event
                        instance.save()

                    for instance in speakers_formset:
                        instance = instance.save(commit=False)
                        instance.event = event
                        instance.save()

                except Exception:
                    if (Event.objects.filter(id=event.id).exists()):
                        Event.objects.filter(id=event.id).delete()
                    return render(request, template_name, {
                        'guests_formset': honoredguests_formset,
                        'subceremonies_formset': subceremony_formset,
                        'speakers_formset': speakers_formset,
                        'form': event_form,
                        'success': False,
                    })
                send_email(request.user.email, request.build_absolute_uri(
                    event.get_absolute_url()))
                return HttpResponseRedirect(reverse("account"))
        else:
            return render(request, template_name, {
                'guests_formset': honoredguests_formset,
                'subceremonies_formset': subceremony_formset,
                'speakers_formset': speakers_formset,
                'form': event_form,
                'success': False,
            })
    return render(request, template_name, {
        'guests_formset': honoredguests_formset,
        'subceremonies_formset': subceremony_formset,
        'speakers_formset': speakers_formset,
        'form': event_form,
        'success': True,
    })


@login_required
def show_event(request, *args, **kwargs):
    template_name = "event_view.html"
    try:
        event = Event.objects.get(id=kwargs['pk'])

        subceremonies = Subceremony.objects.filter(event=event)
        honored_guests = HonoredGuest.objects.filter(event=event)
        speakers = Speaker.objects.filter(event=event)
        if event.user == request.user or request.user.is_superuser or request.user.is_staff:
            return render(request, template_name=template_name, context={'event': event, 'subceremonies': subceremonies, 'honored_guests': honored_guests, 'speakers': speakers})
        else:
            return Http404
    except KeyError:
        return Http404


@login_required
def download_event(request, *args, **kwargs):
    user = request.user
    if user.is_staff or user.is_superuser:
        event_id = kwargs['pk']
        event = Event.objects.get(id=event_id)
        try:
            # path = os.path.join(settings.MEDIA_ROOT,
            #                     "events/event_%s.xlsx" % event_id)
            # f = open(path)
            # f.close()
            path = create_file(event)
        except FileNotFoundError:
            # path = create_file(event)
            return Http404
        response = FileResponse(open(path, 'rb'))
        filename = "event_%s.xlsx" % event_id
        response['Content-Disposition'] = 'inline; filename=' + filename
        return response
    return Http404


def copy_cell(new_cell, cell):
    new_cell.value = copy(cell.value)
    new_cell.font = copy(cell.font)
    new_cell.border = copy(cell.border)
    new_cell.fill = copy(cell.fill)
    new_cell.number_format = copy(cell.number_format)
    new_cell.protection = copy(cell.protection)
    new_cell.alignment = copy(cell.alignment)


def create_file(event):
    colors = [openpyxl.styles.PatternFill(start_color='ffb0a0c6', end_color='ffb0a0c6', fill_type='solid'),
              openpyxl.styles.PatternFill(start_color='ff66ff66',
                                          end_color='ff66ff66', fill_type='solid'),
              openpyxl.styles.PatternFill(start_color='ff0099ff', end_color='ff0099ff', fill_type='solid')]
    header_colors = [openpyxl.styles.PatternFill(start_color='ff60497a', end_color='ff60497a', fill_type='solid'),
                     openpyxl.styles.PatternFill(
                         start_color='ff33cc33', end_color='ff33cc33', fill_type='solid'),
                     openpyxl.styles.PatternFill(start_color='ff0066cc', end_color='ff0066cc', fill_type='solid')]

    subceremonies = Subceremony.objects.filter(event=event)
    honoredguests = HonoredGuest.objects.filter(event=event)
    speakers = Speaker.objects.filter(event=event)
    max_signatories_count = max(
        [subceremony.signatories.count() for subceremony in subceremonies])

    output_path = os.path.join(
        settings.MEDIA_ROOT, 'events/event_%s.xlsx' % event.id)

    doc = openpyxl.Workbook()
    template = openpyxl.load_workbook(os.path.join(
        settings.STATIC_ROOT, "excel/template1.xlsx"))

    sheet = doc.active
    tmpl = template.active

    # copy base info
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    copy_cell(sheet.cell(1, 1), tmpl.cell(1, 1))

    for i in range(1, 7):
        copy_cell(sheet.cell(2, i), tmpl.cell(2, i))

    # copy signatory info cells
    for i in range(max_signatories_count):
        sheet.merge_cells(start_row=1, start_column=7 + i * 5,
                          end_row=1, end_column=7 + i * 5 + 4)
        copy_cell(sheet.cell(1, 7 + i * 5), tmpl.cell(1, 7))
        sheet.cell(1, 7 + i * 5).value = "Сторона подписания %s" % (i + 1)
        sheet.cell(1, 7 + i * 5).fill = header_colors[i % 3]
        for j in range(5):
            copy_cell(sheet.cell(2, 7 + i * 5 + j), tmpl.cell(2, 7 + j))
            sheet.cell(2, 7 + i * 5 + j).fill = colors[i % 3]

    # copy additional and contact info
    for i in range(1, 3):
        for j in range(10):
            copy_cell(sheet.cell(
                i, 7 + (max_signatories_count * 5 + j)), tmpl.cell(i, j + 12))

    # merge cells
    sheet.merge_cells(start_row=1, start_column=7 + max_signatories_count
                      * 5, end_row=1, end_column=7 + max_signatories_count * 5 + 4)
    sheet.merge_cells(start_row=1, start_column=7 + max_signatories_count
                      * 5 + 5, end_row=1, end_column=7 + max_signatories_count * 5 + 9)

    # fill data
    general_data = [event.doc_name, event.window.date,
                    event.window.time, event.window.place, "Церемония подписания соглашения"]
    subceremonies_data = []
    for subceremony in subceremonies:
        subceremony_data = []
        subceremony_data.append(subceremony.signatories.count())
        for signatory in subceremony.signatories.all():
            subceremony_data.append(signatory.org_name)
            subceremony_data.append(" ".join(
                [signatory.signatory_surname, signatory.signatory_name, signatory.signatory_middlename]))
            subceremony_data.append("М" if signatory.is_man else "Ж")
            subceremony_data.append(signatory.position)
            subceremony_data.append(signatory.signatory_name_translate)
        subceremonies_data.append(subceremony_data)

    additional_data = []
    additional_data.append(
        "Да, " + ", ".join([" - ".join([guest.position, guest.name]) for guest in honoredguests]) if honoredguests.count() > 0 else "нет")
    additional_data.append(
        "Да, " + ", ".join([speaker.name for speaker in speakers]) if speakers.count() > 0 else "нет")
    additional_data.append("да" if event.press_briefing else "нет")
    additional_data.append("да" if event.is_online else "нет")
    additional_data.append("да" if event.presents else "нет")

    additional_data.append(
        ", ".join([event.name, event.position, event.org_name]))
    additional_data.append(event.phone_number)
    additional_data.append(event.email)

    additional_data.append(
        ", ".join([event.repr_name, event.repr_position, event.repr_org_name]))
    additional_data.append(event.repr_phone_number)
    additional_data.append(event.repr_email)

    for i in range(1, 6):
        print(general_data[i - 1])
        sheet.cell(3, i).value = general_data[i - 1]
        sheet.merge_cells(start_row=3, start_column=i, end_row=3 +
                          (subceremonies.count() - 1), end_column=i)

    for i in range(len(subceremonies_data)):
        for j in range(len(subceremonies_data[i])):
            sheet.cell(i + 3, j + 6).value = subceremonies_data[i][j]

    for i in range(10):
        sheet.cell(3, 7 + 5 * max_signatories_count +
                   i).value = additional_data[i]
        sheet.merge_cells(start_row=3, start_column=7 + 5 * max_signatories_count + i, end_row=3 +
                          (subceremonies.count() - 1), end_column=7 + 5 * max_signatories_count + i)

    doc.save(output_path)

    return output_path
